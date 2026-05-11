"""Activities for expense-processor (refactored onto ``ntro.subledger``).

``categorise_and_insert_expense`` runs once per submitted row: looks up
the vendor in the bundled vendor_map.json, picks the first matching
category (case-insensitive substring), and INSERTs the row as PENDING
in ``ledgers.expenses`` via ``ntro.subledger.open(name="expenses", ...)``.

``commit_expenses`` runs once at the end: applies any HITL corrections
column-by-column (bespoke UPDATEs — the SDK doesn't have a generic
field-update primitive), then flips status PENDING → APPROVED (or
REJECTED) via ``handle.mark_approved()`` / ``handle.mark_rejected()``,
and returns a per-category summary.

The agent-supplied receipt fields go through
``ExpenseSubmissionFields.model_validate`` — stage-2 boundary types do
the coercion (``ForgivingDate`` for fuzzy dates, etc.). The activity
reads typed values directly; no dict-poking or inline parsers.

N-36: rows whose payload fails strict-type validation (bad ISO
currency, negative gross, vat exceeding gross, etc.) DON'T crash the
activity — they're routed to ``sl.insert_needs_attention(...)`` as
``status='NEEDS_ATTENTION'`` ghost rows with the violation context
captured in the row's ``validation_errors`` JSONB. HITL review
surfaces these alongside PENDING rows; the reviewer fixes the typed
columns and transitions the row through to PENDING / APPROVED.

The GL-handoff (``status=POSTED``) lands in N-27 once
``ntro.subledger.proposals`` + ``gl.post_proposals`` exist. This
activity terminates at APPROVED.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4

from pydantic import ValidationError
from temporalio import activity

import ntro.subledger
from ntro.data import get_data_plane
from ntro.ingest import (
    fetch_submitted_document_by_id,
    insert_submitted_record,
)
from ntro.subledger.types.expenses import canonicalize_tabular_expense_row

from .models import (
    CategoriseInput,
    CommitInput,
    ExpensePeriodResult,
    ExpensePersistedEvent,
    ExpenseSubmissionFields,
    PostedBillSummary,
    PostToGlInput,
    PostToGlResult,
)


_VENDOR_MAP_CACHE: list[dict[str, Any]] | None = None


def _derive_period_yyyy_mm(date_value: Any) -> str:
    if hasattr(date_value, "strftime"):
        try:
            return date_value.strftime("%Y-%m")
        except Exception:
            pass
    return datetime.now(timezone.utc).strftime("%Y-%m")


def _parse_csv_rows(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, Any]] = []
    for row in reader:
        rows.append(dict(row))
    return rows


def _parse_xlsx_rows(data: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - optional dependency guard
        raise RuntimeError("openpyxl is required for XLSX uploads") from exc
    wb = load_workbook(filename=io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in headers_raw]
    out: list[dict[str, Any]] = []
    for r in rows_iter:
        row_obj: dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            row_obj[key] = r[idx] if idx < len(r) else None
        if any(v is not None and str(v).strip() for v in row_obj.values()):
            out.append(row_obj)
    return out


def _parse_tabular_rows(content_type: str, data: bytes) -> list[dict[str, Any]]:
    ct = (content_type or "").lower()
    # Inline base64 uploads (used by e2e/agent paths) can arrive as
    # application/octet-stream. Sniff payload signature and fall back to
    # CSV parsing so these inputs don't get stuck in retry loops.
    if ct == "application/octet-stream":
        if data[:4] == b"PK\x03\x04":
            return _parse_xlsx_rows(data)
        return _parse_csv_rows(data)
    if "csv" in ct or ct == "application/vnd.ms-excel":
        return _parse_csv_rows(data)
    if "sheet" in ct or "excel" in ct:
        return _parse_xlsx_rows(data)
    raise ValueError(f"Unsupported content_type for expense tabular ingest: {content_type}")


@activity.defn(name="expense_processor.ingest_submitted_document")
async def ingest_submitted_document(
    payload: dict[str, Any],
    task_id: str,
) -> list[dict[str, Any]]:
    """Parse an uploaded expense file, persist canonical row events in ingest schema."""
    tenant_slug = str(payload.get("tenant_slug", ""))
    document_ref = str(payload.get("document_ref", ""))
    content_type = str(payload.get("content_type", ""))
    filename = str(payload.get("filename", ""))
    source = str(payload.get("source", "expense-upload"))
    entity_slug = str(payload.get("entity_slug", ""))
    if not tenant_slug or not document_ref or not entity_slug:
        activity.logger.warning("document payload missing required fields: %s", payload)
        return []

    db = await get_data_plane(tenant_slug)
    doc = await fetch_submitted_document_by_id(db, document_ref=document_ref)
    if doc is None:
        activity.logger.warning("document_ref %s not found in submitted_documents", document_ref)
        return []

    resolved_content_type = content_type or doc.content_type or ""
    rows = _parse_tabular_rows(resolved_content_type, doc.data_bytes)

    accepted: list[dict[str, Any]] = []
    now = datetime.now(timezone.utc).isoformat()
    for idx, raw_row in enumerate(rows, start=1):
        canonical, errors = canonicalize_tabular_expense_row(raw_row)
        if errors:
            continue
        event_id = str(uuid4())
        source_ref = {
            "kind": "expense_tabular_row",
            "document_ref": document_ref,
            "file_name": filename or doc.filename,
            "row_index": idx,
            "source": source,
        }
        rec = await insert_submitted_record(
            db,
            id=event_id,
            entity_slug=entity_slug,
            task_id=task_id,
            kind="expense",
            source_ref=source_ref,
            data=canonical,
        )
        accepted.append(
            {
                "event_id": rec.id,
                "kind": "expense",
                "tenant_slug": tenant_slug,
                "entity_slug": entity_slug,
                "source_ref": source_ref,
                "data": canonical,
                "received_at": now,
            }
        )

    activity.logger.info(
        "expense.document.ingested document_ref=%s parsed=%s accepted=%s",
        document_ref,
        len(rows),
        len(accepted),
    )
    return accepted


def _load_vendor_map() -> list[dict[str, Any]]:
    global _VENDOR_MAP_CACHE
    if _VENDOR_MAP_CACHE is None:
        path = Path(__file__).parent / "vendor_map.json"
        raw = json.loads(path.read_text())
        _VENDOR_MAP_CACHE = raw.get("categories", [])
    return _VENDOR_MAP_CACHE


def _propose_category(vendor: str) -> tuple[str | None, str, float]:
    """Return (category, source, confidence). Source is "vendor_lookup"
    when a match was found, "manual" otherwise (reviewer fills in).
    LLM fallback is intentionally out of scope for the PoC — every
    miss surfaces as PENDING with category=null."""
    if not vendor:
        return None, "manual", 0.0
    needle = vendor.lower()
    for entry in _load_vendor_map():
        for match in entry["matches"]:
            if match in needle:
                return entry["category"], "vendor_lookup", 0.85
    return None, "manual", 0.0


def _format_validation_errors(exc: ValidationError) -> list[dict[str, Any]]:
    """Project Pydantic ``ValidationError`` into the row.validation_errors
    JSONB shape: ``[{field, code, message, evaluator}]`` — one entry
    per invariant that failed. Materialises the ``ntro.checks`` "types"
    evaluator surface (see ``ntro.capabilities.checks.QualityCheckSpec``).
    """
    out: list[dict[str, Any]] = []
    for err in exc.errors():
        loc = ".".join(str(part) for part in err.get("loc", []))
        out.append({
            "field": loc or "<unknown>",
            "code": err.get("type", "unknown"),
            "message": err.get("msg", ""),
            "evaluator": "types",
        })
    return out


@activity.defn(name="expense_processor.categorise_and_insert")
async def categorise_and_insert_expense(
    input: CategoriseInput,
) -> ExpensePersistedEvent:
    """Categorise one expense row + INSERT as PENDING (or NEEDS_ATTENTION).

    The agent supplies receipt fields under ``payload.data``. We
    ``model_validate`` them through ``ExpenseSubmissionFields`` —
    boundary types do the coercion (``ForgivingDate`` accepts ISO,
    locale strings, natural language; degrades to ``None`` on garbage).
    Strict types raise (``CurrencyCode`` on a bad ISO; ``Field(gt=0)``
    on negative gross; ``@model_validator`` on vat > gross).

    On strict-validation failure we route the row to
    ``sl.insert_needs_attention(...)`` rather than crashing the
    activity — the row lands as ``status='NEEDS_ATTENTION'`` with the
    raw payload + violation list captured for HITL diagnosis. The
    other rows in a bulk batch are unaffected.
    """
    db = await get_data_plane(input.tenant_slug)
    sl = ntro.subledger.open(
        name="expenses",
        entity_id=input.entity_id,
        task_id=input.task_id,
        db=db,
    )
    raw_payload = input.payload.data

    # Stage 1: validate the submission via ExpenseSubmissionFields.
    # Catch raises here so a single bad row in a 5-row batch doesn't
    # kill the workflow.
    try:
        submission = ExpenseSubmissionFields.model_validate(raw_payload)
    except ValidationError as exc:
        errors = _format_validation_errors(exc)
        # Best-effort partial fields: try to salvage anything that
        # survived the failed validation by best-effort coercion.
        # Vendor + notes are plain str so they almost always parse.
        # event_id is required (NOT NULL on the row); always pass it.
        partial: dict[str, Any] = {"event_id": input.payload.event_id}
        for name in ("vendor", "notes"):
            v = raw_payload.get(name)
            if isinstance(v, str):
                partial[name] = v
        row_id = await sl.insert_needs_attention(
            period=_derive_period_yyyy_mm(raw_payload.get("date")),
            raw_payload=raw_payload,
            validation_errors=errors,
            source_ref=f"event:{input.payload.event_id}",
            type_specific_fields=partial,
        )
        activity.logger.warning(
            "expense.row.needs_attention id=%s event_id=%s violations=%s",
            row_id, input.payload.event_id, [e["field"] for e in errors],
        )
        return ExpensePersistedEvent(
            id=str(row_id),
            event_id=str(input.payload.event_id),
            status="NEEDS_ATTENTION",
            validation_errors=errors,
            vendor=partial.get("vendor"),
            notes=partial.get("notes"),
        )

    vendor = submission.vendor.strip()
    # currency is already upper-normalised by CurrencyCode's BeforeValidator.
    currency = submission.currency

    category, category_source, confidence = _propose_category(vendor)
    row_period = input.period or _derive_period_yyyy_mm(submission.date)

    # Stage 2: build + insert the strict Row. Catches @model_validator
    # failures (vat > gross) — those raise from sl.row(...) since the
    # validator runs at construction time.
    try:
        row = sl.row(
            period=row_period,
            event_id=input.payload.event_id,
            source_ref=f"event:{input.payload.event_id}",
            vendor=vendor,
            amount_gross=submission.amount,
            currency=currency,
            expense_date=submission.date,
            payment_method=submission.payment_method,
            line_items=submission.line_items,
            vat_amount=submission.vat_amount,
            notes=submission.notes,
            category=category,
            category_source=category_source,
            confidence=Decimal(str(confidence)),
        )
    except ValidationError as exc:
        errors = _format_validation_errors(exc)
        partial: dict[str, Any] = {
            "event_id": input.payload.event_id,
            "vendor": vendor,
            "notes": submission.notes,
        }
        row_id = await sl.insert_needs_attention(
            period=row_period,
            raw_payload=raw_payload,
            validation_errors=errors,
            source_ref=f"event:{input.payload.event_id}",
            type_specific_fields={k: v for k, v in partial.items() if v is not None},
        )
        activity.logger.warning(
            "expense.row.needs_attention id=%s event_id=%s violations=%s",
            row_id, input.payload.event_id, [e["field"] for e in errors],
        )
        return ExpensePersistedEvent(
            id=str(row_id),
            event_id=str(input.payload.event_id),
            status="NEEDS_ATTENTION",
            validation_errors=errors,
            vendor=vendor,
            notes=submission.notes,
        )

    await sl.insert(row)

    activity.logger.info(
        "expense.row.persisted id=%s vendor=%s amount=%s category=%s",
        row.id, vendor, submission.amount, category,
    )

    return ExpensePersistedEvent(
        id=str(row.id),
        event_id=str(row.event_id),
        vendor=vendor,
        amount_gross=float(submission.amount),
        currency=currency,
        expense_date=submission.date.isoformat() if submission.date else None,
        payment_method=submission.payment_method,
        line_items=submission.line_items,
        vat_amount=float(submission.vat_amount) if submission.vat_amount is not None else None,
        category=category,
        category_source=category_source,
        confidence=float(confidence),
        status=row.status.value,
        notes=submission.notes,
    )


# Field whitelist for HITL corrections. The SDK doesn't expose a
# generic field-update primitive; corrections are field-name-driven
# (column-by-column UPDATEs) so we keep them inline. Whitelist guards
# against arbitrary column writes from agent-supplied payloads.
_ALLOWED_CORRECTION_FIELDS: set[str] = {
    "vendor",
    "amount_gross",
    "currency",
    "expense_date",
    "payment_method",
    "category",
    "vat_amount",
    "notes",
}


@activity.defn(name="expense_processor.commit_expenses")
async def commit_expenses(input: CommitInput) -> ExpensePeriodResult:
    """Apply HITL corrections, flip statuses, return summary.

    Corrections: bespoke field-name-driven UPDATEs scoped to
    ``ledgers.expenses``. Status flip: ``handle.mark_approved()`` /
    ``handle.mark_rejected()`` from the SDK.

    Summary: category-grouped totals come from a direct SQL aggregate
    rather than the SDK ``query()`` helper — query() returns Row
    instances and we'd group in Python; SQL is simpler for SUM.
    """
    db = await get_data_plane(input.tenant_slug)
    sl = ntro.subledger.open(
        name="expenses",
        entity_id=input.entity_id,
        task_id=input.task_id,
        db=db,
    )

    corrections_applied = 0
    for c in input.corrections:
        if c.field not in _ALLOWED_CORRECTION_FIELDS:
            activity.logger.warning(
                "ignoring correction with unknown field '%s' for expense %s",
                c.field, c.expense_id,
            )
            continue
        value: Any = c.new_value
        if c.field == "amount_gross" or c.field == "vat_amount":
            value = Decimal(str(value)) if value is not None else None
        await db.execute(
            f"UPDATE ledgers.expenses SET {c.field} = $1, "
            f"category_source = CASE WHEN $2 = 'category' THEN 'manual' ELSE category_source END, "
            f"updated_at = NOW() "
            f"WHERE id = $3::uuid AND entity_id = $4",
            value, c.field, c.expense_id, input.entity_id,
        )
        corrections_applied += 1

    if input.expense_ids:
        ids_uuid = [_to_uuid(s) for s in input.expense_ids]
        # mark_approved / mark_rejected return the actual UPDATE row
        # count (via update_status's from_statuses filter). NEEDS_ATTENTION
        # rows in the id list are left alone — they need HITL fix-up
        # to transition to PENDING first.
        if input.approved:
            await sl.mark_approved(ids=ids_uuid)
        else:
            await sl.mark_rejected(ids=ids_uuid)
    else:
        ids_uuid = []

    summary_rows = await db.fetch(
        """
        SELECT COALESCE(category, 'Uncategorised') AS category,
               SUM(amount_gross)::float AS total
        FROM ledgers.expenses
        WHERE entity_id = $1 AND id = ANY($2::uuid[]) AND status = 'APPROVED'
        GROUP BY 1
        """,
        input.entity_id, ids_uuid,
    )
    by_cat = {r["category"]: float(r["total"]) for r in summary_rows}

    # Final-status counts read from the DB (not derived from
    # input.expense_ids). This captures rows that were rejected via the
    # per-row X (✕) action between review and commit — those rows
    # transitioned NEEDS_ATTENTION → REJECTED out-of-band, and the
    # workflow's batch counts need to honour them.
    status_rows = await db.fetch(
        """
        SELECT status, COUNT(*)::int AS c
        FROM ledgers.expenses
        WHERE entity_id = $1 AND id = ANY($2::uuid[])
        GROUP BY status
        """,
        input.entity_id, ids_uuid,
    )
    counts = {r["status"]: r["c"] for r in status_rows}
    approved_count = counts.get("APPROVED", 0)
    rejected_count = counts.get("REJECTED", 0)
    needs_attention_count = counts.get("NEEDS_ATTENTION", 0)

    period_rows = await db.fetch(
        """
        SELECT DISTINCT period
        FROM ledgers.expenses
        WHERE entity_id = $1 AND id = ANY($2::uuid[])
        ORDER BY period ASC
        """,
        input.entity_id, ids_uuid,
    )
    periods = [str(r["period"]) for r in period_rows if r.get("period")]
    period_label = periods[0] if len(periods) == 1 else ("mixed" if periods else (input.period or "unknown"))

    return ExpensePeriodResult(
        period=period_label,
        entity_slug=input.entity_slug,
        rows_total=len(input.expense_ids),
        rows_approved=approved_count,
        rows_rejected=rejected_count,
        rows_needs_attention=needs_attention_count,
        total_amount_by_category=by_cat,
        corrections_applied=corrections_applied,
    )


def _to_uuid(value: Any):
    """Coerce a string / UUID to UUID. Activity inputs deserialise as
    strings via Temporal's converter — accept both."""
    from uuid import UUID
    return value if isinstance(value, UUID) else UUID(str(value))


@activity.defn(name="expense_processor.post_expenses_to_gl")
async def post_expenses_to_gl(input: PostToGlInput) -> PostToGlResult:
    """Post APPROVED expense rows as Bills to the tenant's GL.

    The flow:

    1. Load APPROVED rows for this batch that haven't yet been
       posted (``posted_to_gl=false``).
    2. ``ExpenseRow.propose_for_gl(rows, task_id=...)`` →
       ``list[BillProposal]``. One proposal per row.
    3. ``gl.for_entity(entity_id, tenant_config=...)`` resolves the
       provider — workspace creds from worker env, tenant-scoped state
       from ``tenant_config["gl"]["options"]``.
    4. For each proposal, ``provider.bills.create(proposal,
       external_id=proposal.idempotency_key)``. Sequenced (not
       parallel) for clearer error attribution + rate-limit safety.
    5. ``handle.mark_posted(gl_external_ids={row.id: bill.id, ...})``
       atomically transitions APPROVED → POSTED + denormalises the
       provider id.

    Error handling:

    - ``ValidationFailedError`` / ``IdempotencyConflictError`` /
      ``ConnectionError_`` → row stays APPROVED on the subledger;
      activity captures the error in the result for HITL diagnosis;
      activity returns successfully (the OTHER rows still post).
    - ``TransientProviderError`` → re-raised so Temporal retry
      policy can re-run the activity. Idempotent retries are safe
      because (a) the GL post itself uses ``external_id`` for
      ApiDeck-side dedup, and (b) ``mark_posted`` filters by
      ``status=APPROVED`` so already-POSTED rows are skipped.
    - Any other exception → re-raised to mark the activity FAILED;
      operator triages.

    No-GL gracefully: if the runbook ran without ``tenant_config.gl``
    (tenant hasn't connected a GL), this activity isn't called at
    all — the workflow's ``_step_post_to_gl`` short-circuits.
    """
    from uuid import UUID
    from ntro.capabilities import gl
    from ntro.capabilities.gl.errors import (
        ConnectionError_,
        IdempotencyConflictError,
        TransientProviderError,
        ValidationFailedError,
    )
    from ntro.subledger.types.expenses import ExpenseRow

    db = await get_data_plane(input.tenant_slug)
    sl = ntro.subledger.open(
        name="expenses",
        entity_id=input.entity_id,
        task_id=input.task_id,
        db=db,
    )

    # Load APPROVED rows (the SDK's query() returns Row instances —
    # exactly the shape propose_for_gl wants). We additionally filter
    # out posted_to_gl=true rows in case a partial-success rerun left
    # some rows posted and others not.
    target_ids = {_to_uuid(s) for s in input.expense_ids}
    rows = await sl.query(status="APPROVED")
    rows = [r for r in rows if r.id in target_ids and not getattr(r, "posted_to_gl", False)]
    if not rows:
        return PostToGlResult(posted=0, failed=0, skipped=0)

    provider = await gl.for_entity(input.entity_id, tenant_config=input.tenant_config)
    proposals = ExpenseRow.propose_for_gl(rows, task_id=input.task_id)

    # Index rows by id so we can carry vendor/amount/etc. into the
    # per-bill summary returned to the workflow.
    rows_by_id = {r.id: r for r in rows}

    # Per-row mark_posted (instead of one big batched call after the
    # whole loop) — so if anything in the post-Xero-success path goes
    # wrong on row N, rows 1..N-1 are already POSTED on the subledger
    # and won't be re-posted to Xero on the activity retry. Caught at
    # N-10 Phase 5 smoke: a CHECK violation on POSTED status (missing
    # from migration 004's CHECK list) caused 6 retries × 6 duplicate
    # Xero bills before we caught it.
    #
    # True Xero-side idempotency (via Idempotency-Key HTTP header) is
    # a separate follow-up — see N-XX. Per-row mark_posted is the
    # belt-and-braces local mitigation.

    posted_refs: dict[UUID, str] = {}
    posted_bills: list[PostedBillSummary] = []
    errors: list[dict[str, Any]] = []
    for proposal in proposals:
        row_id = proposal.source_row_ids[0]
        try:
            result = await provider.bills.create(
                proposal,
                external_id=proposal.idempotency_key,
            )
        except (ValidationFailedError, IdempotencyConflictError, ConnectionError_) as exc:
            # Row-level non-retryable — capture, continue. Each error
            # is stored verbatim on the row's posted_journal_ref later
            # (left blank since post failed). The OTHER rows still
            # post; partial-success is the explicit design.
            activity.logger.warning(
                "post_expenses_to_gl: row %s failed (%s): %s",
                row_id, type(exc).__name__, exc,
            )
            errors.append({
                "row_id": str(row_id),
                "idempotency_key": proposal.idempotency_key,
                "error_class": type(exc).__name__,
                "message": str(exc)[:500],
            })
            continue
        except TransientProviderError:
            # Re-raise so Temporal retries the activity. The successful
            # rows that DID post are not re-posted on the next run
            # because mark_posted's from_statuses filter excludes them.
            raise

        # Successful post — capture the provider's id.
        provider_id = getattr(result, "id", None)
        if provider_id is None:
            errors.append({
                "row_id": str(row_id),
                "idempotency_key": proposal.idempotency_key,
                "error_class": "MissingProviderId",
                "message": "ApiDeck returned a successful create without an id",
            })
            continue
        posted_refs[row_id] = str(provider_id)
        # Capture the per-bill detail for the operator's summary screen.
        src = rows_by_id.get(row_id)
        posted_bills.append(
            PostedBillSummary(
                row_id=str(row_id),
                vendor=getattr(src, "vendor", None) if src else None,
                amount_gross=(
                    float(getattr(src, "amount_gross"))
                    if src and getattr(src, "amount_gross", None) is not None
                    else None
                ),
                currency=getattr(src, "currency", None) if src else None,
                expense_date=(
                    getattr(src, "expense_date").isoformat()
                    if src and getattr(src, "expense_date", None) is not None
                    else None
                ),
                posted_journal_ref=str(provider_id),
                idempotency_key=proposal.idempotency_key,
            )
        )
        # Per-row mark_posted — fence the upstream success against any
        # downstream failure on the next iteration. ``mark_posted``
        # uses ``from_statuses=[APPROVED]`` so on retry the row is
        # filtered out (it's now POSTED), preventing re-create on the
        # external system.
        try:
            await sl.mark_posted(gl_external_ids={row_id: str(provider_id)})
        except Exception as mark_exc:
            # If mark_posted fails for ONE row mid-batch, log + add to
            # errors but DON'T re-raise — the bill IS in Xero. Future
            # runs will re-attempt the post unless the operator
            # manually flips the row's status, but that's a far better
            # failure mode than hammering Xero with duplicates.
            activity.logger.error(
                "post_expenses_to_gl: row %s mark_posted FAILED after Xero "
                "wrote bill %s — row stays APPROVED on subledger but bill "
                "exists upstream (manual reconciliation required): %s",
                row_id, provider_id, mark_exc,
            )
            errors.append({
                "row_id": str(row_id),
                "idempotency_key": proposal.idempotency_key,
                "error_class": "MarkPostedFailedAfterCreate",
                "message": f"Xero bill {provider_id} created but local "
                           f"status flip failed: {mark_exc!s}"[:500],
            })

    # Atomic POSTED transition for the successful posts. Idempotent —
    # re-running this on a partially-posted batch only flips rows that
    # are still APPROVED (mark_posted's from_statuses=[APPROVED]).
    # NOTE: mark_posted is now per-row (above, immediately after each
    # successful Xero create) — the post-loop batched call was removed
    # to avoid retry-driven duplicates.

    # Label the destination so the UI can show "2 bills posted to Xero"
    # rather than "2 bills posted to apideck". When service_id is
    # absent the label degrades to "apideck".
    gl_block = (input.tenant_config or {}).get("gl") or {}
    options = gl_block.get("options") or {}
    service_id = options.get("service_id")
    gl_provider = f"apideck:{service_id}" if service_id else "apideck"

    return PostToGlResult(
        posted=len(posted_refs),
        failed=len(errors),
        skipped=0,
        errors=errors,
        posted_journal_refs={str(rid): ref for rid, ref in posted_refs.items()},
        posted_bills=posted_bills,
        gl_provider=gl_provider,
    )
